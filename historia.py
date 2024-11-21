from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import calendar
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# pedir credenciales por consola
username = input("ingrese su usuaro: ")
password = input("ingrese su contraseña: ")

# configuro el webdriver
driver = webdriver.Chrome()


# funcion para esperar que cargue la tabla
def wait_for_loader_to_disappear(driver, timeout=20):
    """
    Espera hasta que el componente loader desaparezca de la pantalla.

    Args:
        driver: El controlador de Selenium.
        timeout: Tiempo máximo para esperar en segundos.
    """
    WebDriverWait(driver, timeout).until(
        EC.invisibility_of_element((By.CLASS_NAME, "http-loader__wrapper"))
    )


# abro la pagina para loguear el usuario
url_login = "https://login.abc.gob.ar/nidp/idff/sso?id=ABC-Form&sid=2&option=credential&sid=2&target=https://menu.abc.gob.ar/"
driver.get(url_login)
# time.sleep(5)


# Busco los campos del formulario
driver.find_element(By.ID, "Ecom_User_ID").send_keys(username)
driver.find_element(By.ID, "Ecom_Password").send_keys(password)

# Busco el boton para enviar el formulario
# boton1 = driver.find_element(By.XPATH, "//a[text()='ENTRAR']")
# boton1.click()

# Ejecuto la funcion para enviar el formulario
driver.execute_script("return imageSubmit();")

# Espero para verificar resultados
# time.sleep(5)

# Click en el boton de Jubilaciones
enlace_jubilaciones = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located(
        (By.XPATH, "//a[.//span[text()='Jubilaciones']]"))
)
enlace_jubilaciones.click()
# Cambia a la última pestaña abierta
driver.switch_to.window(driver.window_handles[-1])


# Despliego el dropdown de Declaracion Jurada
# Esperar hasta que el botón "Declaración Jurada" sea visible
declaracion_jurada_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, "//a[normalize-space(text())='Declaración Jurada']")
    )
)

# Hacer clic en el botón "Declaración Jurada" para desplegar el menú
declaracion_jurada_button.click()

# Esperar hasta que el menú se despliegue (es decir, tenga la clase "open")
li_declaracion_jurada = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located(
        (
            By.XPATH,
            "//li[contains(@class, 'open') and .//a[normalize-space(text())='Declaración Jurada']]",
        )
    )
)

# Buscar el enlace "Servicios" dentro del ul de este li
enlace_servicios = li_declaracion_jurada.find_element(
    By.XPATH, ".//ul[@class='dropdown-menu']//a[normalize-space(text())='Servicios']"
)

# Hacer clic en "Servicios"
enlace_servicios.click()

wait_for_loader_to_disappear(driver)


# Esperar hasta que el botón "100" esté visible y clickeable
boton_100 = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//button[span[text()='25']]"))
)

# Hacer clic en el botón "100"
boton_100.click()


# extraigo los datos de la tabla
def get_table_data(driver):
    # Localizar la tabla por su selector
    table = driver.find_element(
        By.CSS_SELECTOR, "table[ng-table='tableParams']")

    # Obtener todas las filas de la tabla
    rows = table.find_elements(By.TAG_NAME, "tr")

    # Inicializar una lista para almacenar los datos de la tabla
    table_data = []

    # Recorrer las filas de la tabla
    for row in rows:
        # Obtener todas las celdas de la fila
        cells = row.find_elements(By.TAG_NAME, "td")

        # Si la fila tiene celdas (ignorar la fila de encabezado si tiene celdas)
        if cells:
            row_data = [cell.text for cell in cells]
            table_data.append(row_data)

    return table_data


# Función para verificar si hay una página siguiente
def has_next_page(driver):
    try:
        # Verifica si el botón de siguiente página ("next") está presente
        next_button = driver.find_element(By.CSS_SELECTOR, "ul.pagination li a[ng-switch-when='next']")
        
        # Verificamos si el botón 'next' está habilitado o deshabilitado.
        # Si está habilitado, la clase 'disabled' no debería estar presente en el contenedor 'li'
        next_button_parent = next_button.find_element(By.XPATH, '..')
        
        # Si el elemento 'next' está deshabilitado, el padre tendrá la clase 'disabled'
        if "disabled" in next_button_parent.get_attribute("class"):
            return False
        else:
            return True
    except Exception as e:
        print(f"Error al verificar la paginación: {e}")
        return False

# Función para ir a la siguiente página
def go_to_next_page(driver):
    next_button = driver.find_element(
        By.CSS_SELECTOR, "ul.pagination li a[ng-switch-when='next']"
    )
    next_button.click()
    time.sleep(2)  # Espera para que la nueva página cargue


# esperamos que cargue la tabla
wait_for_loader_to_disappear(driver)

# creamos la tabla
table_data = []

# Agregamos los datos de la primera tabla
table_data.extend(get_table_data(driver))

# Si hay mas paginas agregamos los datos de las siguientes tablas
while has_next_page(driver):
    # Ir a la siguiente página
    go_to_next_page(driver)

    # Esperar a que el loader desaparezca y la tabla se cargue
    wait_for_loader_to_disappear(driver)

    # Obtener los datos de la página actual
    table_data.extend(get_table_data(driver))

# Estructura para almacenar los datos organizados por año, mes y los registros
calendar_data = defaultdict(lambda: defaultdict(list))

# Función para analizar las fechas de cada registro
def add_to_calendar(data):
    for item in data:
        secuencia, regimen, revista, enseñanza, cargo, horas, fecha, distrito, organizacion, numero_escuela, etc = item
        
        # Extraemos las fechas
        fecha_inicio, fecha_fin = fecha.split(" al ")
        fecha_inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
        fecha_fin = datetime.strptime(fecha_fin, "%d/%m/%Y")
        
        # Recorremos los meses desde la fecha de inicio hasta la fecha final
        for year in range(fecha_inicio.year, fecha_fin.year + 1):
            for month in range(1, 13):
                if (year == fecha_inicio.year and month >= fecha_inicio.month) or (year > fecha_inicio.year and year < fecha_fin.year) or (year == fecha_fin.year and month <= fecha_fin.month):
                    month_name = calendar.month_name[month]
                    record = f"{secuencia} - {organizacion} {numero_escuela} - {horas if horas != 'sin cargar' else 'sin cargar'}"
                    calendar_data[year][month].append(record)

# Llamamos a la función para llenar el calendario
add_to_calendar(table_data)
# Función para generar el archivo Excel
def create_excel_from_table_data(table_data, filename="calendar_data.xlsx"):
    # Crear un libro de trabajo y una hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario"

    # Escribir encabezados de los meses en la primera fila
    months = [calendar.month_name[i][:3] for i in range(1, 13)]
    for i, month in enumerate(months, start=2):
        cell = ws.cell(row=1, column=i)
        cell.value = month
        cell.alignment = cell.alignment.copy(horizontal="center")

    # Estructura para almacenar los datos organizados por año, mes y los registros
    calendar_data = defaultdict(lambda: defaultdict(list))

    # Función para analizar las fechas de cada registro y organizarlos
    def add_to_calendar(data):
        for item in data:
            secuencia, regimen, revista, enseñanza, cargo, horas, fecha, distrito, organizacion, numero_escuela, etc = item
            
            # Extraemos las fechas
            fecha_inicio, fecha_fin = fecha.split(" al ")
            fecha_inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
            fecha_fin = datetime.strptime(fecha_fin, "%d/%m/%Y")
            
            # Recorremos los meses desde la fecha de inicio hasta la fecha final
            for year in range(fecha_inicio.year, fecha_fin.year + 1):
                for month in range(1, 13):
                    if (year == fecha_inicio.year and month >= fecha_inicio.month) or (year > fecha_inicio.year and year < fecha_fin.year) or (year == fecha_fin.year and month <= fecha_fin.month):
                        month_name = calendar.month_name[month]
                        record = f"{secuencia} - {organizacion} {numero_escuela} - {horas if horas != 'sin cargar' else 'sin cargar'}"
                        calendar_data[year][month].append(record)

    # Llamamos a la función para organizar los datos
    add_to_calendar(table_data)

    # Escribir los datos por año
    row_num = 2  # Empezamos desde la segunda fila (debajo de los meses)
    for year, months_data in calendar_data.items():
        # Escribir el año en la primera columna
        ws.cell(row=row_num, column=1, value=year)
        row_num += 1  # Ir a la siguiente fila después de escribir el año

        # Escribir los registros de cada mes
        max_rows = 0  # Para controlar la cantidad máxima de registros por mes
        for month in range(1, 13):
            column = month + 1  # Los meses empiezan en la columna 2
            if month in months_data:
                records = months_data[month]
                max_rows = max(max_rows, len(records))  # Actualizamos la cantidad máxima de registros por mes
                # Escribir cada registro en una nueva fila debajo del mes correspondiente
                for idx, record in enumerate(records, start=row_num):
                    ws.cell(row=idx, column=column, value=record)

        # Después de cada año, ajustar la fila de inicio para el siguiente año
        row_num += max_rows  # Moverse a la siguiente fila después de escribir todos los registros

        # Dejar espacio en blanco solo después de cada año
        row_num += 1  # Esto asegura que haya una fila vacía entre los años.

    # Ajustar el tamaño de las columnas
    for col in range(1, 14):  # 13 columnas (A - M)
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)  # Añadir un poco de margen
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar el archivo Excel
    wb.save(filename)
    print(f"Archivo {filename} guardado correctamente.")

# Llamamos a la función para generar el Excel con los datos
create_excel_from_table_data(table_data)


time.sleep(9)
driver.quit()
