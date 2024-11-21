from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.metrics import dp
from kivy.clock import Clock
from kivy.core.window import Window
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime
import time
import calendar
import sys

class LoginScreen(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation="vertical", spacing=dp(10), padding=dp(20), **kwargs)

        # Cambiar el color de fondo
        Window.clearcolor = (57 / 255, 167 / 255, 188 / 255, 1)  # Color #39A7BC

        # Centrar los elementos verticalmente
        self.spacing = dp(20)
        self.size_hint = (None, None)
        self.width = dp(300)
        self.height = dp(400)
        self.pos_hint = {"center_x": 0.5, "center_y": 0.5}

        # Título principal
        self.add_widget(Label(
            text="Historia Laboral",
            font_size="24sp",
            bold=True,
            size_hint=(1, None),
            height=dp(50),
            color=(1, 1, 1, 1)  # Blanco
        ))

        # Etiqueta y campo de entrada para el usuario
        self.add_widget(Label(text="Usuario:", size_hint=(1, None), height=dp(30), color=(1, 1, 1, 1)))
        self.username_input = TextInput(
            multiline=False, size_hint=(1, None), height=dp(40), padding=(10, 10)
        )
        self.add_widget(self.username_input)

        # Etiqueta y campo de entrada para la contraseña
        self.add_widget(Label(text="Contraseña:", size_hint=(1, None), height=dp(30), color=(1, 1, 1, 1)))
        self.password_input = TextInput(
            multiline=False, password=True, size_hint=(1, None), height=dp(40), padding=(10, 10)
        )
        self.add_widget(self.password_input)

        # Botón para iniciar el proceso
        self.start_button = Button(
            text="Iniciar proceso",
            size_hint=(None, None),
            width=dp(200),
            height=dp(50),
            pos_hint={"center_x": 0.5},
        )
        self.start_button.bind(on_press=self.start_process)
        self.add_widget(self.start_button)

    def start_process(self, instance):
        username = self.username_input.text
        password = self.password_input.text

        if not username or not password:
            self.show_popup("Error", "Por favor, complete ambos campos.")
            return

        popup = self.show_popup("Información", "El proceso ha comenzado. Esto puede tardar unos minutos.")
        Clock.schedule_once(lambda dt: popup.dismiss(), 3)  # Cerrar automáticamente en 3 segundos

        process_data(username, password)
        Clock.schedule_once(self.finish_process, 5)  # Simula el final del proceso después de 5 segundos

    def finish_process(self, dt):
        popup = self.show_popup("Éxito", "El archivo Excel se generó correctamente.")
        popup.bind(on_dismiss=lambda instance: sys.exit(0))  # Salir de la app al cerrar el popup

    def show_popup(self, title, message):
        popup_content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(20))
        popup_content.add_widget(Label(text=message, size_hint=(1, None), height=dp(40)))
        close_button = Button(text="Cerrar", size_hint=(None, None), width=dp(100), height=dp(40))
        popup = Popup(title=title, content=popup_content, size_hint=(0.8, None), height=dp(200))
        close_button.bind(on_press=popup.dismiss)
        popup_content.add_widget(close_button)
        popup.open()
        return popup



def process_data(username, password):
    # Configurar el WebDriver
    chrome_options = Options()
    chrome_options.add_argument("--incognito")  # Abrir en modo incógnito
    driver = webdriver.Chrome(options=chrome_options)

    # Función para esperar que desaparezca el loader
    def wait_for_loader_to_disappear(driver, timeout=25):
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element((By.CLASS_NAME, "http-loader__wrapper"))
        )

    # Función para obtener los datos de la tabla
    def get_table_data(driver):
        # Localizar la tabla por su selector
        table = driver.find_element(
            By.CSS_SELECTOR, "table[ng-table='tableParams']")
        
        # Obtener todas las filas de la tabla
        rows = table.find_elements(By.TAG_NAME, "tr")

        # Inicializar una lista para almacenar los datos de la tabla
        table_data = []
        time.sleep(2)
        # Recorrer las filas de la tabla
        for row in rows:
            # Obtener todas las celdas de la fila
            cells = row.find_elements(By.TAG_NAME, "td")

            # Si la fila tiene celdas (ignorar la fila de encabezado si tiene celdas)
            if cells:
                row_data = [cell.text for cell in cells]
                table_data.append(row_data)

        return table_data

    # Función para verificar si hay una página siguiente
    def has_next_page(driver):
        try:
            # Verifica si el botón de siguiente página ("next") está presente
            next_button = driver.find_element(By.CSS_SELECTOR, "ul.pagination li a[ng-switch-when='next']")

            # Verificamos si el botón 'next' está habilitado o deshabilitado.
            # Si está habilitado, la clase 'disabled' no debería estar presente en el contenedor 'li'
            next_button_parent = next_button.find_element(By.XPATH, '..')

            # Si el elemento 'next' está deshabilitado, el padre tendrá la clase 'disabled'
            if "disabled" in next_button_parent.get_attribute("class"):
                return False
            else:
                return True
        except Exception as e:
            print(f"Error al verificar la paginación: {e}")
            return False

    # Función para ir a la siguiente página
    def go_to_next_page(driver):
        next_button = driver.find_element(
            By.CSS_SELECTOR, "ul.pagination li a[ng-switch-when='next']"
        )
        next_button.click()
        time.sleep(2)  # Espera para que la nueva página cargue

    # Iniciar sesión
    url_login = "https://login.abc.gob.ar/nidp/idff/sso?id=ABC-Form&sid=2&option=credential&sid=2&target=https://menu.abc.gob.ar/"
    driver.get(url_login)

    driver.find_element(By.ID, "Ecom_User_ID").send_keys(username)
    driver.find_element(By.ID, "Ecom_Password").send_keys(password)
    driver.execute_script("return imageSubmit();")

    # Navegar a la sección deseada
    enlace_jubilaciones = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located(
        (By.XPATH, "//a[.//span[text()='Jubilaciones']]"))
    )
    enlace_jubilaciones.click()
    # Cambia a la última pestaña abierta
    driver.switch_to.window(driver.window_handles[-1])


    # Despliego el dropdown de Declaracion Jurada
    # Esperar hasta que el botón "Declaración Jurada" sea visible
    declaracion_jurada_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, "//a[normalize-space(text())='Declaración Jurada']")
        )
    )

    # Hacer clic en el botón "Declaración Jurada" para desplegar el menú
    declaracion_jurada_button.click()

    # Esperar hasta que el menú se despliegue (es decir, tenga la clase "open")
    li_declaracion_jurada = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(
            (
                By.XPATH,
                "//li[contains(@class, 'open') and .//a[normalize-space(text())='Declaración Jurada']]",
            )
        )
    )

    # Buscar el enlace "Servicios" dentro del ul de este li
    enlace_servicios = li_declaracion_jurada.find_element(
        By.XPATH, ".//ul[@class='dropdown-menu']//a[normalize-space(text())='Servicios']"
    )

    # Hacer clic en "Servicios"
    enlace_servicios.click()
    wait_for_loader_to_disappear(driver)

    # Esperar hasta que el botón "100" esté visible y clickeable
    boton_100 = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[span[text()='25']]"))
    )

    # Hacer clic en el botón "100"
    boton_100.click()

    
    # esperamos que cargue la tabla
    wait_for_loader_to_disappear(driver)

    # creamos la tabla
    table_data = []

    # Agregamos los datos de la primera tabla
    table_data.extend(get_table_data(driver))

    # Si hay mas paginas agregamos los datos de las siguientes tablas
    while has_next_page(driver):
        # Ir a la siguiente página
        go_to_next_page(driver)

        # Esperar a que el loader desaparezca y la tabla se cargue
        wait_for_loader_to_disappear(driver)

        # Obtener los datos de la página actual
        table_data.extend(get_table_data(driver))

    # Función para aplicar una línea horizontal gruesa
    def apply_horizontal_line(ws, row, columns, border_type="thick"):
        side = Side(style=border_type)
        border = Border(bottom=side)
        for col in columns:
            ws.cell(row=row, column=col).border = border


    # Función para generar la hoja organizada por años y meses
    def create_calendar_sheet(wb, table_data):
        ws_calendar = wb.create_sheet(title="Calendario")
        months = [calendar.month_name[i][:3] for i in range(1, 13)]

        # Encabezados de los meses
        for i, month in enumerate(months, start=2):
            ws_calendar.cell(row=1, column=i).value = month

        calendar_data = defaultdict(lambda: defaultdict(list))

        # Llenar el diccionario con los datos
        for item in table_data:
            secuencia, regimen, revista, enseñanza, cargo, horas, fecha, distrito, organizacion, numero_escuela, etc = item
            fecha_inicio, fecha_fin = fecha.split(" al ")
            fecha_inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
            fecha_fin = datetime.strptime(fecha_fin, "%d/%m/%Y")

            for year in range(fecha_inicio.year, fecha_fin.year + 1):
                for month in range(1, 13):
                    if (year == fecha_inicio.year and month >= fecha_inicio.month) or \
                        (year > fecha_inicio.year and year < fecha_fin.year) or \
                        (year == fecha_fin.year and month <= fecha_fin.month):
                        record = f"{secuencia} - {organizacion} {numero_escuela} - {horas if horas != 'sin cargar' else 'sin cargar'}"
                        calendar_data[year][month].append(record)

        # Escribir datos en la hoja
        row_num = 2
        for year, months_data in calendar_data.items():
            # Aplicar línea gruesa antes del año
            if row_num > 2:  # Saltear la primera fila
                apply_horizontal_line(ws_calendar, row_num - 1, range(1, 14))

            # Añadir el año
            ws_calendar.cell(row=row_num, column=1).value = year
            row_num += 1

            max_rows = 0
            for month in range(1, 13):
                column = month + 1
                if month in months_data:
                    records = months_data[month]
                    max_rows = max(max_rows, len(records))
                    for idx, record in enumerate(records, start=row_num):
                        ws_calendar.cell(row=idx, column=column).value = record

            row_num += max_rows

        # Ajustar ancho de columnas
        for col in range(1, 14):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws_calendar.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws_calendar.column_dimensions[column_letter].width = max_length + 2


    # Función para generar la hoja con datos crudos
    def create_raw_data_sheet(wb, table_data):
        ws_raw = wb.create_sheet(title="Datos Originales")
        headers = ["Secuencia", "Regimen", "Revista", "Enseñanza", "Cargo", "Horas", "Fecha", "Distrito", "Organización", "Número Escuela", "Etc"]

        # Escribir cabeceras
        ws_raw.append(headers)

        # Escribir datos
        for item in table_data:
            ws_raw.append(item)

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws_raw.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws_raw.column_dimensions[column_letter].width = max_length + 2


    # Función principal para generar el archivo Excel
    def create_excel_file(table_data, filename="calendar_data.xlsx"):
        wb = Workbook()
        wb.remove(wb.active)  # Quitar la hoja predeterminada

        # Crear ambas hojas
        create_calendar_sheet(wb, table_data)
        create_raw_data_sheet(wb, table_data)

        # Guardar el archivo
        wb.save(filename)
        print(f"Archivo {filename} guardado correctamente.")

    # Llamar a la función con datos de ejemplo
    create_excel_file(table_data)


    driver.quit()


class MyApp(App):
    def build(self):
        return LoginScreen()


if __name__ == "__main__":
    MyApp().run()
