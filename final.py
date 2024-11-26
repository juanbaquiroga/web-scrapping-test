import sys
import time
import calendar
import threading
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from kivy.app import App
from kivy.metrics import dp
from kivy.clock import Clock
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.progressbar import ProgressBar



class LoginScreen(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(orientation="vertical", spacing=dp(10), padding=dp(20), **kwargs)

        # Cambiar el color de fondo
        Window.clearcolor = (57 / 255, 167 / 255, 188 / 255, 1)  # Color #39A7BC

        # Centrar los elementos verticalmente
        self.spacing = dp(20)
        self.size_hint = (None, None)
        self.width = dp(300)
        self.height = dp(400)
        self.pos_hint = {"center_x": 0.5, "center_y": 0.5}

        # Título principal
        self.add_widget(Label(
            text="Historia Laboral",
            font_size="24sp",
            bold=True,
            size_hint=(1, None),
            height=dp(50),
            color=(1, 1, 1, 1)  # Blanco
        ))

        # Etiqueta y campo de entrada para el usuario
        self.add_widget(Label(text="Usuario:", size_hint=(1, None), height=dp(30), color=(1, 1, 1, 1)))
        self.username_input = TextInput(
            multiline=False, size_hint=(1, None), height=dp(40), padding=(10, 10)
        )
        self.add_widget(self.username_input)

        # Etiqueta y campo de entrada para la contraseña
        self.add_widget(Label(text="Contraseña:", size_hint=(1, None), height=dp(30), color=(1, 1, 1, 1)))
        self.password_input = TextInput(
            multiline=False, password=True, size_hint=(1, None), height=dp(40), padding=(10, 10)
        )
        self.add_widget(self.password_input)

        # Botón para iniciar el proceso
        self.start_button = Button(
            text="Iniciar proceso",
            size_hint=(None, None),
            width=dp(200),
            height=dp(50),
            pos_hint={"center_x": 0.5},
        )
        self.start_button.bind(on_press=self.start_process)
        self.add_widget(self.start_button)

    def start_process(self, instance):
        username = self.username_input.text
        password = self.password_input.text

        if not username or not password:
            self.show_popup("Error", "Por favor, complete ambos campos.")
            return

        # Mostrar la pantalla de carga
        self.loading_popup = self.show_loading_popup("Procesando", "Por favor, espere mientras se generan los datos...")
        
        # Ejecutar el proceso en un hilo separado
        threading.Thread(target=self.run_process, args=(username, password)).start()

    def run_process(self, username, password):
        try:
            process_data(username, password)  # Llamar a la función de procesamiento principal
            # Usar Clock para interactuar con la interfaz desde el hilo
            Clock.schedule_once(self.finish_process, 0)
        except Exception as e:
            Clock.schedule_once(lambda dt: self.show_popup("Error", str(e)), 0)
        finally:
            Clock.schedule_once(lambda dt: self.loading_popup.dismiss(), 0)

    def finish_process(self, dt):
        popup = self.show_popup("Éxito", "El archivo Excel se generó correctamente.")
        popup.bind(on_dismiss=lambda instance: sys.exit(0))  # Salir de la app al cerrar el popup

    def show_loading_popup(self, title, message):
        popup_content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(20))
        popup_content.add_widget(Label(text=message, size_hint=(1, None), height=dp(40)))
        progress = ProgressBar(max=1, value=0)
        popup_content.add_widget(progress)
        popup = Popup(title=title, content=popup_content, size_hint=(0.8, None), height=dp(200))
        popup.open()
        return popup

    def show_popup(self, title, message):
        popup_content = BoxLayout(orientation="vertical", spacing=dp(10), padding=dp(20))
        popup_content.add_widget(Label(text=message, size_hint=(1, None), height=dp(40)))
        close_button = Button(text="Cerrar", size_hint=(None, None), width=dp(100), height=dp(40))
        popup = Popup(title=title, content=popup_content, size_hint=(0.8, None), height=dp(200))
        close_button.bind(on_press=popup.dismiss)
        popup_content.add_widget(close_button)
        popup.open()
        return popup

def process_data(username, password):
    # Configurar el WebDriver
    chrome_options = Options()
    chrome_options.add_argument("--incognito")  # Abrir en modo incógnito
    chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(options=chrome_options)

    # Función para esperar que desaparezca el loader
    def wait_for_loader_to_disappear(driver, timeout=25):
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element((By.CLASS_NAME, "http-loader__wrapper"))
        )

    # Función para obtener los datos de la tabla
    def get_table_data(driver):
        # Localizar la tabla por su selector
        table = driver.find_element(
            By.CSS_SELECTOR, "table[ng-table='tableParams']")
        
        # Obtener todas las filas de la tabla
        time.sleep(2)
        rows = table.find_elements(By.TAG_NAME, "tr")

        # Inicializar una lista para almacenar los datos de la tabla
        table_data = []
        # Recorrer las filas de la tabla
        for row in rows:
            # Obtener todas las celdas de la fila
            cells = row.find_elements(By.TAG_NAME, "td")

            # Si la fila tiene celdas (ignorar la fila de encabezado si tiene celdas)
            if cells:
                row_data = [cell.text for cell in cells]
                table_data.append(row_data)

        return table_data

    def ordenar_datos_por_ano_y_mes(table_data):
        """
        Organiza los datos en un diccionario estructurado por año y mes.
        
        Formato de salida:
        {
            año: {
                mes: [
                    [secuencia, organizacion, numero, horas],
                    ...
                ],
                ...
            },
            ...
        }
        """
        datos_ordenados = defaultdict(lambda: defaultdict(list))
        
        for item in table_data:
            # Extraer los campos relevantes
            secuencia, _, _, _, _, horas, fecha, _, organizacion, numero_escuela, _ = item
            
            # Separar la fecha de inicio y fin
            fecha_inicio, fecha_fin = fecha.split(" al ")
            fecha_inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
            fecha_fin = datetime.strptime(fecha_fin, "%d/%m/%Y")
            
            # Iterar entre los años y meses relevantes
            for year in range(fecha_inicio.year, fecha_fin.year + 1):
                for month in range(1, 13):
                    # Validar si el mes y año están dentro del rango de las fechas
                    if (year == fecha_inicio.year and month >= fecha_inicio.month) or \
                    (year > fecha_inicio.year and year < fecha_fin.year) or \
                    (year == fecha_fin.year and month <= fecha_fin.month):
                        # Formar el registro para este mes
                        registro = [secuencia, organizacion, numero_escuela, horas]
                        datos_ordenados[year][month].append(registro)
        
        # Convertir defaultdict a dict para retornarlo
        return {year: dict(months) for year, months in datos_ordenados.items()}

    # Función para verificar si hay una página siguiente
    def has_next_page(driver):
        try:
            # Verifica si el botón de siguiente página ("next") está presente
            next_button = driver.find_element(By.CSS_SELECTOR, "ul.pagination li a[ng-switch-when='next']")

            # Verificamos si el botón 'next' está habilitado o deshabilitado.
            # Si está habilitado, la clase 'disabled' no debería estar presente en el contenedor 'li'
            next_button_parent = next_button.find_element(By.XPATH, '..')

            # Si el elemento 'next' está deshabilitado, el padre tendrá la clase 'disabled'
            if "disabled" in next_button_parent.get_attribute("class"):
                return False
            else:
                return True
        except Exception as e:
            print(f"Error al verificar la paginación: {e}")
            return False

    # Función para ir a la siguiente página
    def go_to_next_page(driver):
        next_button = driver.find_element(
            By.CSS_SELECTOR, "ul.pagination li a[ng-switch-when='next']"
        )
        next_button.click()
        time.sleep(2)  # Espera para que la nueva página cargue

    # Iniciar sesión
    url_login = "https://login.abc.gob.ar/nidp/idff/sso?id=ABC-Form&sid=2&option=credential&sid=2&target=https://menu.abc.gob.ar/"
    driver.get(url_login)

    driver.find_element(By.ID, "Ecom_User_ID").send_keys(username)
    driver.find_element(By.ID, "Ecom_Password").send_keys(password)
    driver.execute_script("return imageSubmit();")

    # Navegar a la sección deseada
    enlace_jubilaciones = WebDriverWait(driver, 10).until(
    EC.presence_of_element_located(
        (By.XPATH, "//a[.//span[text()='Jubilaciones']]"))
    )
    enlace_jubilaciones.click()
    # Cambia a la última pestaña abierta
    driver.switch_to.window(driver.window_handles[-1])


    # Despliego el dropdown de Declaracion Jurada
    # Esperar hasta que el botón "Declaración Jurada" sea visible
    declaracion_jurada_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, "//a[normalize-space(text())='Declaración Jurada']")
        )
    )

    # Hacer clic en el botón "Declaración Jurada" para desplegar el menú
    declaracion_jurada_button.click()

    # Esperar hasta que el menú se despliegue (es decir, tenga la clase "open")
    li_declaracion_jurada = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located(
            (
                By.XPATH,
                "//li[contains(@class, 'open') and .//a[normalize-space(text())='Declaración Jurada']]",
            )
        )
    )

    # Buscar el enlace "Servicios" dentro del ul de este li
    enlace_servicios = li_declaracion_jurada.find_element(
        By.XPATH, ".//ul[@class='dropdown-menu']//a[normalize-space(text())='Servicios']"
    )

    # Hacer clic en "Servicios"
    enlace_servicios.click()
    wait_for_loader_to_disappear(driver)

    # Esperar hasta que el botón "100" esté visible y clickeable
    boton_100 = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[span[text()='100']]"))
    )

    # Hacer clic en el botón "100"
    boton_100.click()

    
    # esperamos que cargue la tabla
    wait_for_loader_to_disappear(driver)

    # creamos la tabla
    table_data = []

    # Agregamos los datos de la primera tabla
    table_data.extend(get_table_data(driver))

    # Si hay mas paginas agregamos los datos de las siguientes tablas
    while has_next_page(driver):
        # Ir a la siguiente página
        go_to_next_page(driver)

        # Esperar a que el loader desaparezca y la tabla se cargue
        wait_for_loader_to_disappear(driver)

        # Obtener los datos de la página actual
        table_data.extend(get_table_data(driver))
    
    # Ordenamos los datos por año y mes.
    calendar_data = ordenar_datos_por_ano_y_mes(table_data)


    # Generar la hoja de Historia Laboral
    def generate_calendar_sheet(workbook, data):
        """
        Genera una hoja de Excel con los años y meses como especificaste.

        Args:
            workbook: El objeto Workbook de openpyxl donde se añadirá la hoja.
            data: Un diccionario con la estructura especificada:
                {
                    año: {
                        mes: [
                            [secuencia, organizacion, numero, horas],
                            ...
                        ],
                        ...
                    },
                    ...
                }
        """
        # Crear una nueva hoja
        ws = workbook.create_sheet("Historia Laboral")
        
        months = [calendar.month_name[i][:3] for i in range(1, 13)]

        # Encabezados de los meses
        for i, month in enumerate(months, start=2):
            ws.cell(row=1, column=i).value = month
        
        row_num = 2  # Comienza en la segunda fila para los datos
        for year, months_data in sorted(data.items()):  # Ordenar los años
            # Insertar el año en la primera columna
            ws.cell(row=row_num, column=1).value = year
            start_row = row_num
            
            # Aplicar línea gruesa si no es la primera fila
            if row_num > 2:
                apply_thick_border(ws, row=row_num - 1, columns=range(1, 14))
            
            # Añadir los datos de cada mes
            max_row_offset = 0  # Para rastrear la cantidad máxima de filas ocupadas por los meses
            for month_idx, month_records in months_data.items():
                col = month_idx + 1  # La columna correspondiente al mes
                for idx, record in enumerate(month_records):
                    # Añadir cada registro debajo del encabezado del mes
                    ws.cell(row=row_num + idx, column=col).value = " - ".join(map(str, record))
                # Rastrea el mayor número de filas que un mes ocupa
                max_row_offset = max(max_row_offset, len(month_records))
            
            # Avanza las filas según el mayor número de registros por mes
            row_num += max_row_offset

            if row_num > start_row:  # Solo si hay registros
                ws.merge_cells(
                    start_row=start_row, start_column=1, end_row=row_num - 1, end_column=1
                )
                year_cell = ws.cell(row=start_row, column=1)
                year_cell.value = year
                year_cell.alignment = Alignment(vertical="top", horizontal="center") 

        # Ajustar el ancho de las columnas
        adjust_column_width(ws)

    def apply_thick_border(worksheet, row, columns):
        """
        Aplica una línea horizontal gruesa en una fila específica.

        Args:
            worksheet: La hoja de trabajo de openpyxl.
            row: La fila donde se aplicará la línea gruesa.
            columns: Un iterable con los índices de las columnas para aplicar la línea.
        """
        thick_border = Border(bottom=Side(style="thick"))
        for col in columns:
            worksheet.cell(row=row, column=col).border = thick_border

    def adjust_column_width(worksheet):
        """
        Ajusta el ancho de las columnas de una hoja en base al contenido.

        Args:
            worksheet: La hoja de trabajo de openpyxl a ajustar.
        """
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 2



    # Función para generar la hoja con datos crudos
    def create_raw_data_sheet(wb, table_data):
        ws_raw = wb.create_sheet(title="Datos Originales")
        headers = ["Secuencia", "Regimen", "Revista", "Enseñanza", "Cargo", "Horas", "Fecha", "Distrito", "Organización", "Número Escuela", "Etc"]

        # Escribir cabeceras
        ws_raw.append(headers)

        # Escribir datos
        for item in table_data:
            ws_raw.append(item)

        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws_raw.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws_raw.column_dimensions[column_letter].width = max_length + 2


    # Función principal para generar el archivo Excel
    def create_excel_file(table_data, filename="calendar_data.xlsx"):
        wb = Workbook()
        wb.remove(wb.active)  # Quitar la hoja predeterminada

        # Crear ambas hojas
        create_raw_data_sheet(wb, table_data)
        generate_calendar_sheet(wb, calendar_data)

        # Guardar el archivo
        wb.save(filename)
        print(f"Archivo {filename} guardado correctamente.")

    # Llamar a la función con datos de ejemplo
    create_excel_file(table_data)
    driver.quit()

class MyApp(App):
    def build(self):
        return LoginScreen()


if __name__ == "__main__":
    MyApp().run()

